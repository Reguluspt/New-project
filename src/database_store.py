from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


HEADER_ROW = 3
DATA_START_ROW = 4

DATABASE_COLUMNS = {
    "stt": 1,
    "customer_info": 2,
    "customer_address": 3,
    "asset_description": 4,
    "valuation_purpose": 5,
    "valuation_fee": 6,
    "valuation_fee_words": 7,
    "personal_note": 8,
    "contract_number": 9,
    "source": 10,
}

_ONES = [
    "không",
    "một",
    "hai",
    "ba",
    "bốn",
    "năm",
    "sáu",
    "bảy",
    "tám",
    "chín",
]


def parse_money(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().lower()
    if not text:
        return None

    multiplier = 1
    if "tỷ" in text:
        multiplier = 1000000000
    elif "triệu" in text or "tr" in text:
        multiplier = 1000000

    # Giữ lại số và dấu chấm/phẩy để xử lý số thập phân (vd: 1.5 triệu)
    clean_text = "".join(ch for ch in text if ch.isdigit() or ch in ".,")
    if not clean_text:
        return None

    try:
        # Thay thế dấu phẩy bằng dấu chấm để chuyển thành float
        val_float = float(clean_text.replace(",", "."))
        return int(val_float * multiplier)
    except (ValueError, TypeError):
        # Fallback: chỉ lấy các chữ số nếu parse float thất bại
        digits = "".join(ch for ch in clean_text if ch.isdigit())
        if not digits:
            return None
        return int(digits) * multiplier


def format_money(value: Any) -> str:
    amount = parse_money(value)
    if amount is None:
        return str(value or "")
    return f"{amount:,}".replace(",", ".")


def _read_three_digits(number: int, *, full: bool) -> str:
    hundreds = number // 100
    tens = (number % 100) // 10
    ones = number % 10
    words: list[str] = []

    if hundreds:
        words.extend([_ONES[hundreds], "trăm"])
    elif full and (tens or ones):
        words.extend(["không", "trăm"])

    if tens > 1:
        words.extend([_ONES[tens], "mươi"])
        if ones == 1:
            words.append("mốt")
        elif ones == 4:
            words.append("tư")
        elif ones == 5:
            words.append("lăm")
        elif ones:
            words.append(_ONES[ones])
    elif tens == 1:
        words.append("mười")
        if ones == 5:
            words.append("lăm")
        elif ones:
            words.append(_ONES[ones])
    elif ones:
        if hundreds or full:
            words.append("lẻ")
        words.append(_ONES[ones])

    return " ".join(words)


def money_to_vietnamese_words(value: Any) -> str:
    amount = parse_money(value)
    if amount is None:
        return ""
    if amount == 0:
        return "Không đồng."

    groups = []
    number = amount
    while number > 0:
        groups.append(number % 1000)
        number //= 1000

    units = ["", "ngàn", "triệu", "tỷ", "ngàn tỷ", "triệu tỷ"]
    parts: list[str] = []
    for idx in range(len(groups) - 1, -1, -1):
        group = groups[idx]
        if group == 0:
            continue
        full = idx < len(groups) - 1
        text = _read_three_digits(group, full=full)
        unit = units[idx] if idx < len(units) else ""
        parts.append(f"{text} {unit}".strip())

    sentence = " ".join(parts).strip()
    sentence = sentence[:1].upper() + sentence[1:]
    return f"{sentence} đồng chẵn."


def _last_data_row(ws: Worksheet) -> int:
    last = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, 11)):
            last = row
    return last


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.protection:
            dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def append_case_to_database(
    database_path: str | Path,
    values: dict[str, Any],
    *,
    sheet_name: str | None = None,
) -> tuple[Path, str, int]:
    path = Path(database_path)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    last_row = _last_data_row(ws)
    target_row = last_row + 1
    source_style_row = last_row if last_row >= DATA_START_ROW else HEADER_ROW
    _copy_row_style(ws, source_style_row, target_row)

    next_stt = 1
    if last_row >= DATA_START_ROW:
        previous_stt = ws.cell(last_row, DATABASE_COLUMNS["stt"]).value
        try:
            next_stt = int(previous_stt) + 1
        except (TypeError, ValueError):
            next_stt = last_row - DATA_START_ROW + 2

    amount = values.get("valuation_fee_number") or values.get("valuation_fee")
    row_values = {
        "stt": next_stt,
        "customer_info": values.get("customer_info", ""),
        "customer_address": values.get("customer_address", ""),
        "asset_description": values.get("asset_description", ""),
        "valuation_purpose": values.get("valuation_purpose", ""),
        "valuation_fee": format_money(amount),
        "valuation_fee_words": values.get("valuation_fee_words") or money_to_vietnamese_words(amount),
        "personal_note": values.get("personal_note", ""),
        "contract_number": values.get("contract_number", ""),
        "source": values.get("source", ""),
    }

    for key, col in DATABASE_COLUMNS.items():
        cell = ws.cell(target_row, col)
        cell.value = row_values[key]
        # Bật wrap_text cho ô chứa ký tự xuống dòng (đa tài sản)
        if isinstance(row_values[key], str) and "\n" in row_values[key]:
            from openpyxl.styles import Alignment
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)
    return path, ws.title, target_row


def read_recent_cases(
    database_path: str | Path,
    *,
    sheet_name: str | None = None,
    limit: int = 8,
) -> list[dict[str, Any]]:
    path = Path(database_path)
    if not path.exists():
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    last_row = _last_data_row(ws)
    rows = []
    for row in range(max(DATA_START_ROW, last_row - limit + 1), last_row + 1):
        rows.append(
            {
                "STT": ws.cell(row, 1).value,
                "Khach hang": ws.cell(row, 2).value,
                "Tai san": ws.cell(row, 4).value,
                "Phi": ws.cell(row, 6).value,
                "So hop dong": ws.cell(row, 9).value,
                "Ngan hang": ws.cell(row, 10).value,
            }
        )
    return rows
