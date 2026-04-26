from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_case_rows_to_excel(
    rows: Sequence[dict[str, object]],
    columns: Sequence[str],
    output_path: str | Path,
    sheet_name: str = "Ho so",
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.freeze_panes = "A2"

    ordered_columns = list(columns)
    sheet.append(ordered_columns)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append([row.get(column, "") for column in ordered_columns])

    for index, column_name in enumerate(ordered_columns, start=1):
        values = [str(column_name)]
        values.extend(str(row.get(column_name, "") or "") for row in rows)
        max_len = max((len(value) for value in values), default=10)
        width = min(max(max_len + 2, 12), 45)
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
