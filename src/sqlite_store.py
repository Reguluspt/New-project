from __future__ import annotations

import re
import json
import sqlite3
import unicodedata
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .database_store import format_money, money_to_vietnamese_words, parse_money


CASE_FIELDS = [
    "customer_type",
    "case_status",
    "cancel_reason",
    "execution_month",
    "payment_status",
    "contract_number",
    "contract_date",
    "certificate_date",
    "asset_type",
    "asset_description",
    "preliminary_status",
    "expected_finish_date",
    "valuation_purpose",
    "source",
    "customer_info",
    "customer_phone",
    "customer_address",
    "citizen_id",
    "valuation_fee_number",
    "valuation_fee_words",
    "advance_payment",
    "survey_cost",
    "business_staff",
    "valuation_staff",
    "controller",
    "legal_note",
    "personal_note",
    "so_thua_dat",
    "so_to_ban_do",
    "dia_chi_thua_dat",
    "owner_name",
    "case_folder",
    "original_file_path",
    "tax_code",
    "representative_name",
    "representative_position",
    "authorization_note",
    "handover_contact_name",
    "handover_contact_position",
    "handover_contact_phone",
    "valuation_branch",
    "office",
    "web_case_id",
    "certificate_number",
    "delivery_contact_id",
    "tracking_number",
]

OPTIONAL_COLUMNS = {
    "customer_type": "TEXT DEFAULT 'individual'",
    "case_status": "TEXT DEFAULT 'Đang xử lý'",
    "cancel_reason": "TEXT",
    "execution_month": "TEXT",
    "payment_status": "TEXT DEFAULT 'Đã thanh toán'",
    "contract_date": "TEXT",
    "customer_phone": "TEXT",
    "certificate_date": "TEXT",
    "case_folder": "TEXT",
    "original_file_path": "TEXT",
    "tax_code": "TEXT",
    "representative_name": "TEXT",
    "representative_position": "TEXT",
    "authorization_note": "TEXT",
    "handover_contact_name": "TEXT",
    "handover_contact_position": "TEXT",
    "handover_contact_phone": "TEXT",
    "valuation_branch": "TEXT DEFAULT 'cn Đà Nẵng'",
    "office": "TEXT DEFAULT 'vp Đà Nẵng'",
    "web_case_id": "TEXT",
    "certificate_number": "TEXT",
    "delivery_contact_id": "INTEGER",
    "tracking_number": "TEXT",
}

DEFAULT_PAYMENT_STATUS = "Đã thanh toán"
DEFAULT_CASE_STATUS = "Đang xử lý"
CANCELED_CASE_STATUS = "Hủy"
DEFAULT_EXECUTION_MONTH = datetime.now().strftime("%m/%Y")
MONTH_SORT_EXPR = """
CASE
    WHEN execution_month GLOB '[0-1][0-9]/[0-9][0-9][0-9][0-9]'
    THEN SUBSTR(execution_month, 4, 4) || SUBSTR(execution_month, 1, 2)
    ELSE ''
END
"""
SORTABLE_FIELDS = {
    "created_at": "created_at",
    "execution_month": MONTH_SORT_EXPR,
    "valuation_fee_number": "COALESCE(valuation_fee_number, 0)",
    "customer_info": "customer_info COLLATE NOCASE",
    "contract_number": "contract_number COLLATE NOCASE",
    "payment_status": "payment_status COLLATE NOCASE",
    "case_status": "case_status COLLATE NOCASE",
    "source": "source COLLATE NOCASE",
}

ORGANIZATION_KEYWORDS = [
    "CONG TY",
    "CÔNG TY",
    "TNHH",
    "CO PHAN",
    "CỔ PHẦN",
    "NGAN HANG",
    "NGÂN HÀNG",
    "CHI NHANH",
    "CHI NHÁNH",
    "DOANH NGHIEP",
    "DOANH NGHIỆP",
    "HTX",
    "HOP TAC XA",
    "HỢP TÁC XÃ",
]

HEADER_ALIASES = {
    "contract_number": ["so hop dong", "dien giai"],
    "customer_info": ["khach hang"],
    "customer_address": ["dia chi"],
    "asset_description": ["tai san tham dinh"],
    "valuation_purpose": ["muc dich tham dinh"],
    "valuation_fee_number": ["phi tham dinh"],
    "valuation_fee_words": ["so tien bang chu"],
    "survey_cost": ["cpks + ctv", "cpks"],
    "business_staff": ["nvkd"],
    "source": ["ngan hang", "nguon"],
    "personal_note": ["ghi chu"],
}

REQUIRED_IMPORT_FIELDS = {
    "customer_info",
    "customer_address",
    "asset_description",
    "valuation_purpose",
    "valuation_fee_number",
}


def infer_customer_type(customer_info: str) -> str:
    text = (customer_info or "").upper()
    return "organization" if any(keyword in text for keyword in ORGANIZATION_KEYWORDS) else "individual"


@contextmanager
def connect(db_path: str | Path) -> Iterator[sqlite3.Connection]:
    path = Path(db_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.create_function("CASEFOLD", 1, lambda value: str(value or "").casefold(), deterministic=True)
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


_DB_INITIALIZED = {}

def init_db(db_path: str | Path) -> None:
    path_key = str(db_path)
    if _DB_INITIALIZED.get(path_key):
        return
        
    with connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                customer_type TEXT DEFAULT 'individual',
                case_status TEXT DEFAULT 'Đang xử lý',
                execution_month TEXT,
                payment_status TEXT DEFAULT 'Đã thanh toán',
                contract_number TEXT,
                contract_date TEXT,
                certificate_date TEXT,
                asset_type TEXT,
                asset_description TEXT,
                preliminary_status TEXT,
                expected_finish_date TEXT,
                valuation_purpose TEXT,
                source TEXT,
                customer_info TEXT,
                customer_address TEXT,
                citizen_id TEXT,
                valuation_fee_number INTEGER,
                valuation_fee_words TEXT,
                advance_payment TEXT,
                survey_cost TEXT,
                business_staff TEXT,
                valuation_staff TEXT,
                controller TEXT,
                legal_note TEXT,
                personal_note TEXT,
                so_thua_dat TEXT,
                so_to_ban_do TEXT,
                dia_chi_thua_dat TEXT,
                owner_name TEXT,
                case_folder TEXT,
                original_file_path TEXT,
                tax_code TEXT,
                representative_name TEXT,
                representative_position TEXT,
                authorization_note TEXT,
                handover_contact_name TEXT,
                handover_contact_position TEXT,
                handover_contact_phone TEXT,
                web_case_id TEXT
            )
            """
        )
        existing_columns = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(cases)").fetchall()
        }
        for column, definition in OPTIONAL_COLUMNS.items():
            if column not in existing_columns:
                conn.execute(f"ALTER TABLE cases ADD COLUMN {column} {definition}")
        conn.execute(
            """
            UPDATE cases
            SET case_status = ?
            WHERE case_status IS NULL OR TRIM(case_status) = ''
            """,
            (DEFAULT_CASE_STATUS,),
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_contract ON cases(contract_number)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_customer ON cases(customer_info)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_citizen ON cases(citizen_id)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS case_gcn_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                case_id INTEGER NOT NULL,
                source_file_id TEXT,
                source_file_name TEXT,
                asset_index INTEGER NOT NULL DEFAULT 0,
                so_thua_dat TEXT,
                so_to_ban_do TEXT,
                dia_chi_thua_dat TEXT,
                owner_name TEXT,
                owner_address TEXT,
                owner_citizen_id TEXT,
                so_giay_chung_nhan TEXT,
                so_vao_so_cap_giay_chung_nhan TEXT,
                ngay_cap_giay_chung_nhan TEXT,
                extraction_json TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(case_id) REFERENCES cases(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_case_gcn_details_case ON case_gcn_details(case_id, asset_index)")

        # Create organizations table
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS organizations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tax_code TEXT,
                name TEXT NOT NULL,
                abbreviation TEXT,
                address TEXT,
                representative TEXT,
                position TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_organizations_tax_code ON organizations(tax_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_organizations_name ON organizations(name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_organizations_abbreviation ON organizations(abbreviation)")
        
    _DB_INITIALIZED[path_key] = True


def get_all_organizations(db_path: str | Path) -> list[dict[str, Any]]:
    init_db(db_path)
    with connect(db_path) as conn:
        cursor = conn.execute("SELECT * FROM organizations ORDER BY name ASC")
        return [dict(row) for row in cursor.fetchall()]


def find_organization_by_query(db_path: str | Path, query: str) -> list[dict[str, Any]]:
    search = str(query or "").strip().casefold()
    if not search:
        return []
    with connect(db_path) as conn:
        cursor = conn.execute(
            """
            SELECT * FROM organizations
            WHERE CASEFOLD(name) LIKE :search
               OR CASEFOLD(abbreviation) LIKE :search
               OR CASEFOLD(tax_code) LIKE :search
            """,
            {"search": f"%{search}%"},
        )
        return [dict(row) for row in cursor.fetchall()]


def add_organization(db_path: str | Path, data: dict[str, Any]) -> int:
    now = datetime.now().isoformat()
    with connect(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO organizations (
                tax_code, name, abbreviation, address, representative, position, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data.get("tax_code", ""),
                data.get("name", ""),
                data.get("abbreviation", ""),
                data.get("address", ""),
                data.get("representative", ""),
                data.get("position", ""),
                now,
                now,
            ),
        )
        return cursor.lastrowid


def update_organization(db_path: str | Path, org_id: int, data: dict[str, Any]) -> None:
    now = datetime.now().isoformat()
    with connect(db_path) as conn:
        conn.execute(
            """
            UPDATE organizations
            SET tax_code = ?, name = ?, abbreviation = ?, address = ?, representative = ?, position = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                data.get("tax_code", ""),
                data.get("name", ""),
                data.get("abbreviation", ""),
                data.get("address", ""),
                data.get("representative", ""),
                data.get("position", ""),
                now,
                org_id,
            ),
        )


def delete_organization(db_path: str | Path, org_id: int) -> None:
    with connect(db_path) as conn:
        conn.execute("DELETE FROM organizations WHERE id = ?", (org_id,))


def format_contract_number(value: str) -> str:
    if not value:
        return value
    val = value.strip()
    now = datetime.now()
    current_year = now.strftime("%Y")
    current_month = now.strftime("%m")
    
    if re.match(r"^\d{4}$", val):
        return f"010/{current_year}/N{current_month}-{val}/DN"
    if re.match(r"^\.\d{4}$", val):
        return f"010/{current_year}/N{current_month}.{val[1:]}/DN"
        
    # Handle semi-short format e.g. "N05-0806", "N05.0806", "N05-0851/DN"
    match = re.match(r"^N(\d{2})([-.]\d{4})(/DN)?$", val, re.IGNORECASE)
    if match:
        month, suffix, _ = match.groups()
        return f"010/{current_year}/N{month}{suffix}/DN"
        
    return value


def _normalize_case(values: dict[str, Any]) -> dict[str, Any]:
    normalized = {field: values.get(field, "") for field in CASE_FIELDS}
    normalized["contract_number"] = format_contract_number(str(normalized.get("contract_number") or "").strip())
    normalized["customer_type"] = normalized.get("customer_type") or infer_customer_type(normalized.get("customer_info", ""))
    normalized["case_status"] = str(normalized.get("case_status") or DEFAULT_CASE_STATUS).strip() or DEFAULT_CASE_STATUS
    normalized["cancel_reason"] = str(normalized.get("cancel_reason") or "").strip()
    if normalized["case_status"] != CANCELED_CASE_STATUS:
        normalized["cancel_reason"] = ""
    normalized["execution_month"] = str(normalized.get("execution_month") or DEFAULT_EXECUTION_MONTH).strip() or DEFAULT_EXECUTION_MONTH
    normalized["payment_status"] = str(normalized.get("payment_status") or DEFAULT_PAYMENT_STATUS).strip() or DEFAULT_PAYMENT_STATUS
    fee = parse_money(normalized.get("valuation_fee_number"))
    normalized["valuation_fee_number"] = fee
    normalized["valuation_fee_words"] = normalized.get("valuation_fee_words") or money_to_vietnamese_words(fee)
    if not normalized.get("owner_name"):
        normalized["owner_name"] = normalized.get("customer_info", "")
    return normalized


def _gcn_value(detail: dict[str, Any], *fields: str) -> str:
    for field in fields:
        value = detail.get(field)
        if isinstance(value, dict):
            value = value.get("value", "")
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _insert_case_gcn_details(conn: sqlite3.Connection, case_id: int, gcn_details: list[dict[str, Any]], now: str) -> None:
    for index, detail in enumerate(gcn_details):
        if not isinstance(detail, dict):
            continue
        conn.execute(
            """
            INSERT INTO case_gcn_details (
                case_id, source_file_id, source_file_name, asset_index,
                so_thua_dat, so_to_ban_do, dia_chi_thua_dat,
                owner_name, owner_address, owner_citizen_id,
                so_giay_chung_nhan, so_vao_so_cap_giay_chung_nhan, ngay_cap_giay_chung_nhan,
                extraction_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                case_id,
                str(detail.get("source_file_id") or ""),
                str(detail.get("source_file_name") or ""),
                int(detail.get("asset_index") or index),
                _gcn_value(detail, "so_thua_dat", "so_thua"),
                _gcn_value(detail, "so_to_ban_do", "so_to"),
                _gcn_value(detail, "dia_chi_thua_dat", "land_address"),
                _gcn_value(detail, "ten_chu_so_huu_cuoi_cung", "owner_name"),
                _gcn_value(detail, "dia_chi_chu_so_huu_cuoi_cung", "owner_address"),
                _gcn_value(detail, "so_cccd_chu_so_huu_cuoi_cung", "owner_citizen_id", "so_cccd"),
                _gcn_value(detail, "so_giay_chung_nhan"),
                _gcn_value(detail, "so_vao_so_cap_giay_chung_nhan"),
                _gcn_value(detail, "ngay_cap_giay_chung_nhan"),
                json.dumps(detail, ensure_ascii=False),
                now,
            ),
        )


def create_case(db_path: str | Path, values: dict[str, Any], *, gcn_details: list[dict[str, Any]] | None = None) -> int:
    init_db(db_path)
    now = datetime.now().isoformat(timespec="seconds")
    data = _normalize_case(values)
    columns = ["created_at", "updated_at", *CASE_FIELDS]
    params = {"created_at": now, "updated_at": now, **data}
    placeholders = ", ".join(f":{column}" for column in columns)
    with connect(db_path) as conn:
        cursor = conn.execute(
            f"INSERT INTO cases ({', '.join(columns)}) VALUES ({placeholders})",
            params,
        )
        case_id = int(cursor.lastrowid)
        if gcn_details:
            _insert_case_gcn_details(conn, case_id, gcn_details, now)
        return case_id


def update_case(db_path: str | Path, case_id: int, values: dict[str, Any]) -> None:
    init_db(db_path)
    with connect(db_path) as conn:
        row = conn.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone()
        if row is None:
            raise ValueError(f"Khong tim thay ho so id={case_id}")
        existing = dict(row)
        
        merged = {field: existing.get(field, "") for field in CASE_FIELDS}
        merged.update(values)
        data = _normalize_case(merged)
        data["updated_at"] = datetime.now().isoformat(timespec="seconds")
        data["id"] = case_id
        
        assignments = ", ".join(f"{field} = :{field}" for field in [*CASE_FIELDS, "updated_at"])
        cursor = conn.execute(f"UPDATE cases SET {assignments} WHERE id = :id", data)
        if cursor.rowcount == 0:
            raise ValueError(f"Khong tim thay ho so id={case_id}")


def delete_case(db_path: str | Path, case_id: int) -> None:
    init_db(db_path)
    with connect(db_path) as conn:
        conn.execute("DELETE FROM case_gcn_details WHERE case_id = ?", (case_id,))
        cursor = conn.execute("DELETE FROM cases WHERE id = ?", (case_id,))
        if cursor.rowcount == 0:
            raise ValueError(f"Khong tim thay ho so id={case_id}")


def get_case(db_path: str | Path, case_id: int) -> dict[str, Any] | None:
    init_db(db_path)
    with connect(db_path) as conn:
        row = conn.execute("SELECT * FROM cases WHERE id = ?", (case_id,)).fetchone()
        if row is None:
            return None
        case_data = dict(row)
        case_data["gcn_details"] = [
            dict(detail)
            for detail in conn.execute(
                "SELECT * FROM case_gcn_details WHERE case_id = ? ORDER BY asset_index, id",
                (case_id,),
            ).fetchall()
        ]
    return case_data


def _build_case_search(
    query: str,
    *,
    note_query: str = "",
    case_status: str = "",
    execution_month: str = "",
    payment_status: str = "",
    source: str = "",
    customer_type: str = "",
    business_staff: str = "",
) -> tuple[str, dict[str, Any]]:
    search = query.strip().casefold()
    conditions: list[str] = []
    params: dict[str, Any] = {}
    if query.strip():
        conditions.append(
            """
            (
                INSTR(CASEFOLD(contract_number), :search) > 0
                OR INSTR(CASEFOLD(customer_info), :search) > 0
                OR INSTR(CASEFOLD(customer_address), :search) > 0
                OR INSTR(CASEFOLD(asset_description), :search) > 0
                OR INSTR(CASEFOLD(valuation_purpose), :search) > 0
                OR INSTR(CASEFOLD(source), :search) > 0
                OR INSTR(CASEFOLD(citizen_id), :search) > 0
                OR INSTR(CASEFOLD(personal_note), :search) > 0
                OR INSTR(CASEFOLD(customer_type), :search) > 0
                OR INSTR(CASEFOLD(case_status), :search) > 0
                OR INSTR(CASEFOLD(execution_month), :search) > 0
                OR INSTR(CASEFOLD(payment_status), :search) > 0
            )
            """
        )
        params["search"] = search
    if note_query.strip():
        conditions.append("INSTR(CASEFOLD(personal_note), :note_search) > 0")
        params["note_search"] = note_query.strip().casefold()
    if case_status.strip():
        conditions.append("case_status = :case_status")
        params["case_status"] = case_status.strip()
    if execution_month.strip():
        conditions.append("execution_month = :execution_month")
        params["execution_month"] = execution_month.strip()
    if payment_status.strip():
        conditions.append("payment_status = :payment_status")
        params["payment_status"] = payment_status.strip()
    if source.strip():
        conditions.append("source = :source")
        params["source"] = source.strip()
    if customer_type.strip():
        conditions.append("customer_type = :customer_type")
        params["customer_type"] = customer_type.strip()
    if business_staff.strip():
        conditions.append("business_staff = :business_staff")
        params["business_staff"] = business_staff.strip()

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    return where, params


def count_cases(
    db_path: str | Path,
    query: str = "",
    *,
    note_query: str = "",
    case_status: str = "",
    execution_month: str = "",
    payment_status: str = "",
    source: str = "",
    customer_type: str = "",
    business_staff: str = "",
) -> int:
    init_db(db_path)
    where, params = _build_case_search(
        query,
        note_query=note_query,
        case_status=case_status,
        execution_month=execution_month,
        payment_status=payment_status,
        source=source,
        customer_type=customer_type,
        business_staff=business_staff,
    )
    sql = f"SELECT COUNT(*) FROM cases {where}"
    with connect(db_path) as conn:
        value = conn.execute(sql, params).fetchone()[0]
    return int(value or 0)


def search_cases(
    db_path: str | Path,
    query: str = "",
    *,
    note_query: str = "",
    case_status: str = "",
    execution_month: str = "",
    payment_status: str = "",
    source: str = "",
    customer_type: str = "",
    business_staff: str = "",
    sort_field: str = "created_at",
    sort_direction: str = "desc",
    limit: int = 50,
    offset: int = 0,
) -> list[dict[str, Any]]:
    init_db(db_path)
    where, params = _build_case_search(
        query,
        note_query=note_query,
        case_status=case_status,
        execution_month=execution_month,
        payment_status=payment_status,
        source=source,
        customer_type=customer_type,
        business_staff=business_staff,
    )

    order_field = SORTABLE_FIELDS.get(sort_field, SORTABLE_FIELDS["created_at"])
    order_direction = "ASC" if str(sort_direction).lower() == "asc" else "DESC"

    sql = f"""
        SELECT *
        FROM cases
        {where}
        ORDER BY {order_field} {order_direction}, id DESC
        LIMIT :limit
        OFFSET :offset
    """
    params.update({"limit": limit, "offset": offset})
    with connect(db_path) as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(row) for row in rows]


def recent_cases(db_path: str | Path, *, limit: int = 8) -> list[dict[str, Any]]:
    return search_cases(db_path, "", limit=limit)


def distinct_case_values(db_path: str | Path, field: str) -> list[str]:
    allowed = {
        "execution_month",
        "payment_status",
        "customer_type",
        "case_status",
        "source",
        "business_staff",
        "asset_type",
        "valuation_branch",
    }
    if field not in allowed:
        raise ValueError(f"Unsupported distinct field: {field}")
    init_db(db_path)
    if field == "execution_month":
        sql = f"""
            SELECT DISTINCT execution_month
            FROM cases
            WHERE execution_month IS NOT NULL AND TRIM(execution_month) <> ''
            ORDER BY {MONTH_SORT_EXPR} DESC
        """
    else:
        sql = f"""
            SELECT DISTINCT {field}
            FROM cases
            WHERE {field} IS NOT NULL AND TRIM({field}) <> ''
            ORDER BY {field}
        """
    with connect(db_path) as conn:
        rows = conn.execute(sql).fetchall()
    return [str(row[0]) for row in rows if row[0] not in (None, "")]


def revenue_summary(
    db_path: str | Path,
    *,
    target_month: str,
    query: str = "",
    note_query: str = "",
    case_status: str = "",
    payment_status: str = "",
    source: str = "",
    customer_type: str = "",
    business_staff: str = "",
) -> dict[str, Any]:
    init_db(db_path)
    where, params = _build_case_search(
        query,
        note_query=note_query,
        case_status=case_status,
        payment_status=payment_status,
        source=source,
        customer_type=customer_type,
        business_staff=business_staff,
    )
    base_where = where
    current_month_where = f"{base_where} {'AND' if base_where else 'WHERE'} execution_month = :target_month AND COALESCE(case_status, '') <> :canceled_status"
    cumulative_where = f"{base_where} {'AND' if base_where else 'WHERE'} {MONTH_SORT_EXPR} <= :target_month_sort AND COALESCE(case_status, '') <> :canceled_status"
    previous_where = f"{base_where} {'AND' if base_where else 'WHERE'} execution_month <> :target_month AND {MONTH_SORT_EXPR} < :target_month_sort AND COALESCE(case_status, '') <> :canceled_status"

    params["target_month"] = target_month
    params["target_month_sort"] = _execution_month_sort_value(target_month)
    params["canceled_status"] = CANCELED_CASE_STATUS

    with connect(db_path) as conn:
        case_count_current_month = conn.execute(
            f"SELECT COUNT(*) FROM cases {current_month_where}",
            params,
        ).fetchone()[0]
        projected_current_month = conn.execute(
            f"SELECT COALESCE(SUM(COALESCE(valuation_fee_number, 0)), 0) FROM cases {current_month_where}",
            params,
        ).fetchone()[0]
        paid_current_month = conn.execute(
            f"""
            SELECT COALESCE(SUM(COALESCE(valuation_fee_number, 0)), 0)
            FROM cases
            {current_month_where} AND payment_status = :paid_status
            """,
            {**params, "paid_status": DEFAULT_PAYMENT_STATUS},
        ).fetchone()[0]
        paid_to_date = conn.execute(
            f"""
            SELECT COALESCE(SUM(COALESCE(valuation_fee_number, 0)), 0)
            FROM cases
            {cumulative_where} AND payment_status = :paid_status
            """,
            {**params, "paid_status": DEFAULT_PAYMENT_STATUS},
        ).fetchone()[0]
        previous_rows = conn.execute(
            f"""
            SELECT
                execution_month,
                COALESCE(SUM(COALESCE(valuation_fee_number, 0)), 0) AS projected_revenue,
                COALESCE(SUM(CASE WHEN payment_status = :paid_status THEN COALESCE(valuation_fee_number, 0) ELSE 0 END), 0) AS paid_revenue,
                COUNT(*) AS case_count
            FROM cases
            {previous_where}
            GROUP BY execution_month
            ORDER BY {MONTH_SORT_EXPR} DESC
            """,
            {**params, "paid_status": DEFAULT_PAYMENT_STATUS},
        ).fetchall()

    monthly_rows = []
    for row in previous_rows:
        projected = float(row["projected_revenue"] or 0)
        paid = float(row["paid_revenue"] or 0)
        monthly_rows.append(
            {
                "Tháng": row["execution_month"],
                "Số hồ sơ": int(row["case_count"] or 0),
                "Doanh thu dự kiến": projected,
                "Đã thanh toán": paid,
                "Chưa thanh toán": max(projected - paid, 0),
            }
        )

    projected_current = float(projected_current_month or 0)
    paid_current = float(paid_current_month or 0)
    return {
        "target_month": target_month,
        "case_count_current_month": int(case_count_current_month or 0),
        "projected_current_month": projected_current,
        "paid_current_month": paid_current,
        "unpaid_current_month": max(projected_current - paid_current, 0),
        "paid_to_date": float(paid_to_date or 0),
        "previous_months": monthly_rows,
    }


def _execution_month_sort_value(value: str) -> str:
    text = str(value or "").strip()
    match = re.fullmatch(r"(\d{2})/(\d{4})", text)
    if not match:
        return ""
    return f"{match.group(2)}{match.group(1)}"


def monthly_revenue_breakdown(
    db_path: str | Path,
    *,
    year: str = "",
    query: str = "",
    note_query: str = "",
    case_status: str = "",
    source: str = "",
    customer_type: str = "",
    business_staff: str = "",
) -> list[dict[str, Any]]:
    init_db(db_path)
    where, params = _build_case_search(
        query,
        note_query=note_query,
        case_status=case_status,
        source=source,
        customer_type=customer_type,
        business_staff=business_staff,
    )
    conditions: list[str] = []
    if where:
        conditions.append(where.removeprefix("WHERE ").strip())
    conditions.append("execution_month IS NOT NULL")
    conditions.append("TRIM(execution_month) <> ''")
    conditions.append("COALESCE(case_status, '') <> :canceled_status")
    params["canceled_status"] = CANCELED_CASE_STATUS
    if str(year).strip():
        conditions.append("SUBSTR(execution_month, 4, 4) = :year")
        params["year"] = str(year).strip()
    final_where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    sql = f"""
        SELECT
            execution_month,
            COUNT(*) AS case_count,
            COALESCE(SUM(COALESCE(valuation_fee_number, 0)), 0) AS projected_revenue,
            COALESCE(SUM(CASE WHEN payment_status = :paid_status THEN COALESCE(valuation_fee_number, 0) ELSE 0 END), 0) AS paid_revenue,
            COALESCE(SUM(CASE WHEN payment_status <> :paid_status THEN COALESCE(valuation_fee_number, 0) ELSE 0 END), 0) AS unpaid_revenue
        FROM cases
        {final_where}
        GROUP BY execution_month
        ORDER BY {MONTH_SORT_EXPR} ASC
    """
    with connect(db_path) as conn:
        rows = conn.execute(sql, {**params, "paid_status": DEFAULT_PAYMENT_STATUS}).fetchall()

    results: list[dict[str, Any]] = []
    for row in rows:
        results.append(
            {
                "month": row["execution_month"],
                "case_count": int(row["case_count"] or 0),
                "projected_revenue": float(row["projected_revenue"] or 0),
                "paid_revenue": float(row["paid_revenue"] or 0),
                "unpaid_revenue": float(row["unpaid_revenue"] or 0),
            }
        )
    return results


def display_cases(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    display = []
    for row in rows:
        def _format_multiline(text: Any) -> str:
            text_str = str(text or "")
            if "\n" in text_str:
                return "\n".join(f"- {line.strip()}" for line in text_str.split("\n") if line.strip())
            return text_str

        display.append(
            {
                "ID": row.get("id") or "",
                "Ngày tạo": row.get("created_at") or "",
                "Loại khách hàng": "Cá nhân" if row.get("customer_type") == "individual" else "Tổ chức",
                "Trạng thái hồ sơ": row.get("case_status") or DEFAULT_CASE_STATUS,
                "Tháng thực hiện": row.get("execution_month") or "",
                "Trạng thái thanh toán": row.get("payment_status") or DEFAULT_PAYMENT_STATUS,
                "Số hợp đồng": row.get("contract_number") or "",
                "Ngày hợp đồng": row.get("contract_date") or "",
                "Ngày chứng thư": row.get("certificate_date") or "",
                "Thông tin khách hàng": row.get("customer_info") or "",
                "Địa chỉ khách hàng": row.get("customer_address") or "",
                "CCCD": row.get("citizen_id"),
                "Nguồn/ngân hàng": row.get("source") or "",
                "Phí thẩm định": format_money(row.get("valuation_fee_number")),
                "Sơ bộ": row.get("preliminary_status") or "",
                "Số thửa đất": _format_multiline(row.get("so_thua_dat")),
                "Số tờ bản đồ": _format_multiline(row.get("so_to_ban_do")),
                "Địa chỉ thửa đất": _format_multiline(row.get("dia_chi_thua_dat")),
                "Chủ sở hữu cuối cùng": _format_multiline(row.get("owner_name")),
                "Loại tài sản": row.get("asset_type") or "",
                "Thời gian dự kiến hoàn thành": row.get("expected_finish_date") or "",
                "Tạm ứng": row.get("advance_payment") or "",
                "Tài sản thẩm định giá": _format_multiline(row.get("asset_description")),
                "Mục đích thẩm định": row.get("valuation_purpose") or "",
                "Ghi chú cá nhân": row.get("personal_note") or "",
                "Chi phí khảo sát": row.get("survey_cost") or "",
                "Chuyên viên kinh doanh": row.get("business_staff") or "",
                "Chuyên viên nghiệp vụ": row.get("valuation_staff") or "",
                "Kiểm soát": row.get("controller") or "",
                "Liên hệ lấy pháp lý": row.get("legal_note") or "",
                "Mã số thuế": row.get("tax_code") or "",
                "Người đại diện": row.get("representative_name") or "",
                "Chức vụ người đại diện": row.get("representative_position") or "",
                "Căn cứ/ủy quyền đại diện": row.get("authorization_note") or "",
                "Người nhận bàn giao": row.get("handover_contact_name") or "",
                "Chức vụ người nhận bàn giao": row.get("handover_contact_position") or "",
                "Điện thoại người nhận bàn giao": row.get("handover_contact_phone") or "",
                "Thư mục hồ sơ": row.get("case_folder") or "",
                "File gốc": row.get("original_file_path") or "",
            }
        )
    return display


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("đ", "d").replace("Đ", "D")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _detect_header_row(ws: Worksheet) -> int | None:
    best_row = None
    best_score = 0
    # Kiểm tra 10 dòng đầu tiên thay vì 6 để tăng khả năng tìm thấy header
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        normalized_cells = [_normalize_header(cell.value) for cell in ws[row_idx]]
        score = 0
        for aliases in HEADER_ALIASES.values():
            if any(any(alias in header for alias in aliases) for header in normalized_cells):
                score += 1
        if score > best_score:
            best_row = row_idx
            best_score = score
    return best_row if best_row is not None and best_score >= 4 else None


def _build_sheet_mapping(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    headers = [_normalize_header(cell.value) for cell in ws[header_row]]
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            for index, header in enumerate(headers, start=1):
                if alias in header:
                    mapping[field] = index
                    break
            if field in mapping:
                break
    return mapping


def _get_cell_value(ws: Worksheet, row: int, mapping: dict[str, int], field: str) -> Any:
    column = mapping.get(field)
    if not column:
        return ""
    return ws.cell(row, column).value


def _is_empty_import_row(ws: Worksheet, row: int, mapping: dict[str, int]) -> bool:
    for field in REQUIRED_IMPORT_FIELDS:
        if _get_cell_value(ws, row, mapping, field) not in (None, ""):
            return False
    return True


def _looks_like_contract_number(value: Any) -> bool:
    text = str(value or "").strip().upper()
    if not text:
        return False
    # Chấp nhận các chuỗi có ít nhất 1 chữ số và chứa dấu gạch chéo hoặc gạch ngang
    # Hoặc chuỗi bắt đầu bằng chữ và có số (ví dụ: V05-123)
    return bool(
        re.search(r"\d+.*[/\\-].*\d+", text)
        or re.search(r"[A-Z]+\d+", text)
    )


def _parse_execution_month(sheet_name: str, contract_number: str) -> str:
    normalized_name = _normalize_header(sheet_name)
    match = re.search(r"thang\s+(\d{1,2})\s*[./-]\s*(\d{4})", normalized_name)
    if match:
        return f"{int(match.group(1)):02d}/{int(match.group(2))}"

    contract = str(contract_number or "").strip().upper()
    match = re.search(r"/(\d{4})/[A-Z](\d{2})[-.]", contract)
    if match:
        return f"{int(match.group(2)):02d}/{int(match.group(1))}"

    match = re.search(r"[A-Z](\d{2})(\d{2})[-.]", contract)
    if match:
        return f"{int(match.group(2)):02d}/{2000 + int(match.group(1))}"

    return DEFAULT_EXECUTION_MONTH


def _case_exists(
    db_path: str | Path,
    *,
    contract_number: str,
    execution_month: str,
    customer_info: str,
    asset_description: str,
    valuation_fee_number: Any,
) -> bool:
    with connect(db_path) as conn:
        if contract_number:
            exists = conn.execute(
                "SELECT 1 FROM cases WHERE contract_number = ? LIMIT 1",
                (contract_number,),
            ).fetchone()
            if exists:
                return True
        exists = conn.execute(
            """
            SELECT 1
            FROM cases
            WHERE execution_month = ?
              AND customer_info = ?
              AND asset_description = ?
              AND COALESCE(valuation_fee_number, 0) = COALESCE(?, 0)
            LIMIT 1
            """,
            (
                execution_month,
                customer_info,
                asset_description,
                parse_money(valuation_fee_number),
            ),
        ).fetchone()
    return exists is not None


def _load_import_workbook(excel_path: str | Path):
    path = Path(excel_path)
    return load_workbook(path, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")


def _close_import_workbook(workbook) -> None:
    workbook.close()
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        vba_archive.close()


def list_importable_excel_sheets(excel_path: str | Path) -> list[str]:
    workbook = _load_import_workbook(excel_path)
    try:
        sheet_names: list[str] = []
        for ws in workbook.worksheets:
            header_row = _detect_header_row(ws)
            if not header_row:
                continue
            mapping = _build_sheet_mapping(ws, header_row)
            if REQUIRED_IMPORT_FIELDS.issubset(mapping):
                sheet_names.append(ws.title)
        return sheet_names
    finally:
        _close_import_workbook(workbook)


def import_excel_database(
    db_path: str | Path,
    excel_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> int:
    init_db(db_path)
    workbook = _load_import_workbook(excel_path)
    try:
        worksheets = [workbook[sheet_name]] if sheet_name and sheet_name in workbook.sheetnames else list(workbook.worksheets)

        imported = 0
        for ws in worksheets:
            header_row = _detect_header_row(ws)
            if not header_row:
                continue
            mapping = _build_sheet_mapping(ws, header_row)
            if not REQUIRED_IMPORT_FIELDS.issubset(mapping):
                continue

            for row in range(header_row + 1, ws.max_row + 1):
                if _is_empty_import_row(ws, row, mapping):
                    continue

                raw_contract = _get_cell_value(ws, row, mapping, "contract_number")
                contract_number = str(raw_contract or "").strip()
                if not _looks_like_contract_number(contract_number):
                    contract_number = ""

                customer_info = str(_get_cell_value(ws, row, mapping, "customer_info") or "").strip()
                customer_address = str(_get_cell_value(ws, row, mapping, "customer_address") or "").strip()
                asset_description = str(_get_cell_value(ws, row, mapping, "asset_description") or "").strip()
                valuation_purpose = str(_get_cell_value(ws, row, mapping, "valuation_purpose") or "").strip()
                valuation_fee_number = parse_money(_get_cell_value(ws, row, mapping, "valuation_fee_number"))
                valuation_fee_words = str(_get_cell_value(ws, row, mapping, "valuation_fee_words") or "").strip()
                survey_cost = parse_money(_get_cell_value(ws, row, mapping, "survey_cost"))
                business_staff = str(_get_cell_value(ws, row, mapping, "business_staff") or "").strip()
                source = str(_get_cell_value(ws, row, mapping, "source") or "").strip()
                personal_note = str(_get_cell_value(ws, row, mapping, "personal_note") or "").strip()
                execution_month = _parse_execution_month(ws.title, contract_number)

                if _case_exists(
                    db_path,
                    contract_number=contract_number,
                    execution_month=execution_month,
                    customer_info=customer_info,
                    asset_description=asset_description,
                    valuation_fee_number=valuation_fee_number,
                ):
                    continue

                create_case(
                    db_path,
                    {
                        "execution_month": execution_month,
                        "payment_status": DEFAULT_PAYMENT_STATUS,
                        "customer_info": customer_info,
                        "customer_address": customer_address,
                        "asset_description": asset_description,
                        "valuation_purpose": valuation_purpose,
                        "valuation_fee_number": valuation_fee_number,
                        "valuation_fee_words": valuation_fee_words,
                        "survey_cost": survey_cost or "",
                        "business_staff": business_staff,
                        "personal_note": personal_note,
                        "contract_number": contract_number,
                        "source": source,
                        "customer_type": infer_customer_type(customer_info),
                    },
                )
                imported += 1

        return imported
    finally:
        _close_import_workbook(workbook)
