from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname, range_boundaries


FIELD_CELLS = {
    "contract_number": "D2",
    "asset_type": "D4",
    "asset_description": "D5",
    "preliminary_status": "D6",
    "expected_finish_date": "D7",
    "valuation_purpose": "D9",
    "source": "D10",
    "customer_info": "D11",
    "valuation_fee_text": "D13",
    "valuation_fee_number": "E13",
    "advance_payment": "D15",
    "survey_cost": "D16",
    "business_staff": "D18",
    "valuation_staff": "D19",
    "controller": "D20",
    "other_note": "D21",
    "legal_note": "D22",
    "customer_address": "D23",
    "citizen_id": "D24",
    "personal_note": "D25",
}

EXCEL_DROPDOWN_CELLS = {
    "asset_type": "D4",
    "preliminary_status": "D6",
    "valuation_purpose": "D9",
    "source": "D10",
    "business_staff": "E17",
    "valuation_staff": "D19",
}

EXCEL_DROPDOWN_LABELS = {
    "asset_type": "Loại tài sản",
    "preliminary_status": "Sơ bộ",
    "valuation_purpose": "Mục đích thẩm định",
    "source": "Nguồn/đối tác",
    "business_staff": "Chuyên viên kinh doanh",
    "valuation_staff": "Chuyên viên nghiệp vụ",
}


def compose_asset_description(so_thua: str, so_to: str, dia_chi: str) -> str:
    parts: list[str] = []
    if so_thua:
        parts.append(f"Thua dat so {so_thua}")
    if so_to:
        parts.append(f"to ban do so {so_to}")

    summary = ", ".join(parts)
    if dia_chi:
        if summary:
            return f"{summary}; tai dia chi {dia_chi}."
        return f"Thua dat tai dia chi {dia_chi}."
    return f"{summary}." if summary else ""


def safe_int(value: str) -> int | str:
    cleaned = (value or "").replace(".", "").replace(",", "").replace(" ", "")
    if cleaned.isdigit():
        return int(cleaned)
    return value


def resolve_write_cell(worksheet, cell_ref: str) -> str:
    for merged_range in worksheet.merged_cells.ranges:
        if cell_ref in merged_range:
            return f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
    return cell_ref


def _values_from_range(worksheet, range_ref: str) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref.replace("$", ""))
    values: list[str] = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            value = str(cell.value or "").strip()
            if not value or value in seen:
                continue
            seen.add(value)
            values.append(value)
    return values


def load_dropdown_options(template_path: str | Path, sheet_name: str = "Up Hs") -> dict[str, list[str]]:
    workbook = load_workbook(Path(template_path), data_only=True)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    options: dict[str, list[str]] = {}
    validations = list(worksheet.data_validations.dataValidation) if worksheet.data_validations else []

    for field_name, cell_ref in EXCEL_DROPDOWN_CELLS.items():
        field_options: list[str] = []
        for validation in validations:
            if validation.type != "list" or cell_ref not in validation.sqref:
                continue
            formula = str(validation.formula1 or "").strip()
            if not formula:
                continue
            if formula.startswith('"') and formula.endswith('"'):
                field_options = [item.strip() for item in formula.strip('"').split(",") if item.strip()]
                break
            range_ref = formula.replace("$", "")
            target_sheet = worksheet
            if "!" in range_ref:
                sheet_part, range_part = range_ref.split("!", 1)
                target_sheet = workbook[sheet_part.strip("'")]
                range_ref = range_part
            field_options = _values_from_range(target_sheet, range_ref)
            break
        options[field_name] = field_options
    return options


def _find_dropdown_validation(worksheet, cell_ref: str):
    validations = list(worksheet.data_validations.dataValidation) if worksheet.data_validations else []
    for validation in validations:
        if validation.type == "list" and cell_ref in validation.sqref:
            return validation
    return None


def _dropdown_source_range(workbook, worksheet, field_name: str):
    cell_ref = EXCEL_DROPDOWN_CELLS[field_name]
    validation = _find_dropdown_validation(worksheet, cell_ref)
    if validation is None:
        raise ValueError(f"Không tìm thấy data validation cho ô {cell_ref}")

    formula = str(validation.formula1 or "").strip()
    if not formula or (formula.startswith('"') and formula.endswith('"')):
        raise ValueError(f"Data validation của ô {cell_ref} không dùng vùng dữ liệu có thể sửa")

    range_ref = formula.replace("$", "")
    target_sheet = worksheet
    if "!" in range_ref:
        sheet_part, range_part = range_ref.split("!", 1)
        target_sheet = workbook[sheet_part.strip("'")]
        range_ref = range_part

    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    if min_col != max_col:
        raise ValueError(f"Vùng danh sách của {cell_ref} phải là một cột")
    return validation, target_sheet, min_col, min_row, max_row


def save_dropdown_options(
    template_path: str | Path,
    field_name: str,
    options: list[str],
    sheet_name: str = "Up Hs",
) -> Path:
    if field_name not in EXCEL_DROPDOWN_CELLS:
        raise ValueError(f"Trường danh sách chọn không hợp lệ: {field_name}")

    cleaned: list[str] = []
    seen: set[str] = set()
    for option in options:
        value = str(option or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        cleaned.append(value)

    template = Path(template_path)
    workbook = load_workbook(template)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    validation, source_sheet, column, start_row, end_row = _dropdown_source_range(workbook, worksheet, field_name)

    new_end_row = max(end_row, start_row + max(len(cleaned), 1) - 1)
    for row in range(start_row, new_end_row + 1):
        source_sheet.cell(row=row, column=column).value = None
    for offset, value in enumerate(cleaned):
        source_sheet.cell(row=start_row + offset, column=column).value = value

    list_end_row = start_row + max(len(cleaned), 1) - 1
    column_letter = get_column_letter(column)
    formula_sheet = quote_sheetname(source_sheet.title)
    validation.formula1 = f"{formula_sheet}!${column_letter}${start_row}:${column_letter}${list_end_row}"

    workbook.save(template)
    return template


def fill_template(
    template_path: str | Path,
    output_path: str | Path,
    values: dict[str, Any],
) -> Path:
    template = Path(template_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template)
    ws = wb["Up Hs"] if "Up Hs" in wb.sheetnames else wb.active

    for key, cell in FIELD_CELLS.items():
        if key not in values:
            continue
        value = values[key]
        if value is None:
            continue
        if key == "valuation_fee_number":
            value = safe_int(str(value))
        ws[resolve_write_cell(ws, cell)] = value

    wb.save(output)
    return output
