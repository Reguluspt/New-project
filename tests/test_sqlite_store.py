from __future__ import annotations

import tempfile
import unittest
from zipfile import ZIP_DEFLATED, ZipFile
from pathlib import Path

from openpyxl import Workbook

from src.sqlite_store import (
    CANCELED_CASE_STATUS,
    DEFAULT_CASE_STATUS,
    DEFAULT_EXECUTION_MONTH,
    DEFAULT_PAYMENT_STATUS,
    create_case,
    delete_case,
    get_case,
    import_excel_database,
    init_db,
    list_importable_excel_sheets,
    search_cases,
    update_case,
)


def _append_import_headers_and_row(sheet, *, contract: str, customer: str, fee: int) -> None:
    headers = [
        "Dien giai",
        "Khach hang",
        "Dia chi",
        "Tai san tham dinh",
        "Muc dich tham dinh",
        "Phi tham dinh",
        "So tien bang chu",
        "CPKS",
        "NVKD",
        "Nguon",
        "Ghi chu",
    ]
    sheet.append(headers)
    sheet.append(
        [
            contract,
            customer,
            "Dia chi KH",
            "Tai san",
            "Vay von",
            fee,
            "",
            100000,
            "NVKD 1",
            "VCB",
            "Note",
        ]
    )


def _add_vba_archive_to_xlsm(path: Path) -> None:
    temp_path = path.with_name(f"{path.stem}_with_vba{path.suffix}")
    with ZipFile(path, "r") as source, ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = text.replace(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
                if "vbaProject.bin" not in text:
                    text = text.replace(
                        "</Types>",
                        '<Override PartName="/xl/vbaProject.bin" '
                        'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
                    )
                data = text.encode("utf-8")
            target.writestr(item, data)
        target.writestr("xl/vbaProject.bin", b"fake-vba-project-for-import-test")
    temp_path.replace(path)


class SQLiteStoreCrudTests(unittest.TestCase):
    def test_create_case_applies_defaults_and_normalizes_money(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "cases.db"
            init_db(db_path)

            case_id = create_case(
                db_path,
                {
                    "contract_number": "HD-001",
                    "contract_date": "25/04/2026",
                    "certificate_date": "26/04/2026",
                    "customer_info": "Nguyen Van A",
                    "valuation_fee_number": "1.500.000",
                },
            )
            row = get_case(db_path, case_id)

        self.assertIsNotNone(row)
        assert row is not None
        self.assertEqual(row["contract_number"], "HD-001")
        self.assertEqual(row["contract_date"], "25/04/2026")
        self.assertEqual(row["certificate_date"], "26/04/2026")
        self.assertEqual(row["case_status"], DEFAULT_CASE_STATUS)
        self.assertEqual(row["payment_status"], DEFAULT_PAYMENT_STATUS)
        self.assertEqual(row["execution_month"], DEFAULT_EXECUTION_MONTH)
        self.assertEqual(row["valuation_fee_number"], 1500000)
        self.assertEqual(row["owner_name"], "Nguyen Van A")

    def test_update_case_preserves_cancel_reason_only_for_cancelled_case(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "cases.db"
            init_db(db_path)
            case_id = create_case(db_path, {"customer_info": "Khach A"})

            update_case(
                db_path,
                case_id,
                {
                    "case_status": CANCELED_CASE_STATUS,
                    "cancel_reason": "Khach huy yeu cau",
                    "payment_status": "Chưa thanh toán",
                },
            )
            cancelled = get_case(db_path, case_id)
            update_case(db_path, case_id, {"case_status": DEFAULT_CASE_STATUS})
            restored = get_case(db_path, case_id)

        self.assertIsNotNone(cancelled)
        self.assertIsNotNone(restored)
        assert cancelled is not None
        assert restored is not None
        self.assertEqual(cancelled["case_status"], CANCELED_CASE_STATUS)
        self.assertEqual(cancelled["cancel_reason"], "Khach huy yeu cau")
        self.assertEqual(cancelled["payment_status"], "Chưa thanh toán")
        self.assertEqual(restored["case_status"], DEFAULT_CASE_STATUS)
        self.assertEqual(restored["cancel_reason"], "")

    def test_update_case_persists_web_case_id(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "cases.db"
            init_db(db_path)
            case_id = create_case(db_path, {"customer_info": "Khach A"})

            update_case(db_path, case_id, {"web_case_id": "TS1: 217265 - Dia chi A\nTS2: 217266 - Dia chi B"})
            row = get_case(db_path, case_id)

        self.assertIsNotNone(row)
        assert row is not None
        self.assertEqual(row["web_case_id"], "TS1: 217265 - Dia chi A\nTS2: 217266 - Dia chi B")

    def test_delete_case_removes_record_and_rejects_missing_id(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "cases.db"
            init_db(db_path)
            case_id = create_case(db_path, {"customer_info": "Khach A"})

            delete_case(db_path, case_id)
            deleted = get_case(db_path, case_id)

            with self.assertRaises(ValueError):
                delete_case(db_path, case_id)

        self.assertIsNone(deleted)

    def test_import_excel_database_reads_multiple_importable_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            db_path = root / "cases.db"
            excel_path = root / "multi_sheet.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "Thang 03.2026"
            second = workbook.create_sheet("Thang 04.2026")
            ignored = workbook.create_sheet("Ghi chu")

            for sheet, contract, customer, fee in [
                (first, "N2503-001", "Cong ty ABC", 1000000),
                (second, "N2504-002", "Nguyen Van A", 2000000),
            ]:
                _append_import_headers_and_row(sheet, contract=contract, customer=customer, fee=fee)
            ignored.append(["Khong phai", "du lieu"])
            workbook.save(excel_path)
            workbook.close()

            importable_sheets = list_importable_excel_sheets(excel_path)
            imported = import_excel_database(db_path, excel_path)
            duplicate_imported = import_excel_database(db_path, excel_path)
            rows = search_cases(db_path, "", sort_field="contract_number", sort_direction="asc", limit=10, offset=0)

        self.assertEqual(importable_sheets, ["Thang 03.2026", "Thang 04.2026"])
        self.assertEqual(imported, 2)
        self.assertEqual(duplicate_imported, 0)
        self.assertEqual([row["contract_number"] for row in rows], ["N2503-001", "N2504-002"])
        self.assertEqual([row["execution_month"] for row in rows], ["03/2026", "04/2026"])
        self.assertEqual([row["valuation_fee_number"] for row in rows], [1000000, 2000000])

    def test_import_excel_database_reads_macro_enabled_xlsm(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            db_path = root / "cases.db"
            excel_path = root / "legacy_tracking.xlsm"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Thang 03.2026"
            _append_import_headers_and_row(sheet, contract="N2603-011", customer="Cong ty Macro", fee=3000000)
            workbook.save(excel_path)
            workbook.close()
            _add_vba_archive_to_xlsm(excel_path)

            with ZipFile(excel_path) as archive:
                self.assertIn("xl/vbaProject.bin", archive.namelist())

            importable_sheets = list_importable_excel_sheets(excel_path)
            imported = import_excel_database(db_path, excel_path)
            rows = search_cases(db_path, "", limit=10, offset=0)

        self.assertEqual(importable_sheets, ["Thang 03.2026"])
        self.assertEqual(imported, 1)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["contract_number"], "N2603-011")
        self.assertEqual(rows[0]["execution_month"], "03/2026")
        self.assertEqual(rows[0]["valuation_fee_number"], 3000000)


if __name__ == "__main__":
    unittest.main()
