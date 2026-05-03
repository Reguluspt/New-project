from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook

from src.app_config import UNPAID_STATUS, extraction_to_form_state
from src.case_files import case_folder, save_original_file
from src.excel_writer import fill_template
from src.models import ExtractedValue, LandCertificateExtraction, PageOrientationMetadata
from src.sqlite_store import DEFAULT_CASE_STATUS, create_case, get_case, init_db
from views.entry_actions import apply_gemini_page_metadata_to_viewer, run_ocr_extraction


def _value(value: str, confidence: float = 0.95) -> ExtractedValue:
    return ExtractedValue(value=value, confidence=confidence, evidence=f"evidence {value}")


def _sample_extraction() -> LandCertificateExtraction:
    return LandCertificateExtraction(
        so_thua_dat=_value("123"),
        so_to_ban_do=_value("45"),
        dia_chi_thua_dat=_value("Phuong Tan Loi, TP Buon Ma Thuot"),
        ten_chu_so_huu_cuoi_cung=_value("Nguyen Van A"),
        dia_chi_chu_so_huu_cuoi_cung=_value("12 Le Loi"),
        so_cccd_chu_so_huu_cuoi_cung=_value("012345678901"),
        notes=["OCR toan bo tai lieu"],
    )


def _run_mock_gemini_ocr(extraction: LandCertificateExtraction) -> tuple[LandCertificateExtraction, dict[str, str], Mock]:
    form_state: dict[str, str] = {}
    remember_ai_config = Mock()
    with (
        patch("views.entry_actions.extract_land_certificate_with_gemini", return_value=extraction),
        patch(
            "views.entry_actions.apply_extraction_to_form",
            side_effect=lambda result: form_state.update(extraction_to_form_state(result)),
        ),
    ):
        result = run_ocr_extraction(
            preview_path=Path("full_document.pdf"),
            ai_provider="Gemini",
            api_key="secret",
            model="gemini-test",
            remember_ai_config=remember_ai_config,
        )
    return result, form_state, remember_ai_config


def _create_excel_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Up Hs"
    worksheet["A1"] = "Template"
    workbook.save(path)
    workbook.close()


class EntryActionsTests(unittest.TestCase):
    def test_run_ocr_extraction_uses_gemini_and_applies_form(self) -> None:
        extraction = object()
        remember_ai_config = Mock()
        with (
            patch("views.entry_actions.extract_land_certificate_with_gemini", return_value=extraction) as gemini,
            patch("views.entry_actions.extract_land_certificate") as openai,
            patch("views.entry_actions.apply_extraction_to_form") as apply_form,
        ):
            result = run_ocr_extraction(
                preview_path=Path("gcn.pdf"),
                ai_provider="Gemini",
                api_key="secret",
                model="gemini-test",
                remember_ai_config=remember_ai_config,
            )

        self.assertIs(result, extraction)
        gemini.assert_called_once_with(Path("gcn.pdf"), api_key="secret", model="gemini-test")
        openai.assert_not_called()
        apply_form.assert_called_once_with(extraction)
        remember_ai_config.assert_called_once()

    def test_apply_gemini_page_metadata_updates_viewer_rotations(self) -> None:
        extraction = _sample_extraction().model_copy(
            update={
                "page_metadata": [
                    PageOrientationMetadata(page_number=1, rotation=90),
                    PageOrientationMetadata(page_number=2, rotation=180),
                    PageOrientationMetadata(page_number=3, rotation=45),
                ]
            }
        )
        state: dict[str, object] = {}
        with patch("views.entry_actions.st.session_state", state):
            updated = apply_gemini_page_metadata_to_viewer(Path("gcn.pdf"), extraction)
            rotations = dict(state["page_rotations"])

        self.assertTrue(updated)
        self.assertEqual(rotations[f"{Path('gcn.pdf').resolve()}::1"], 90)
        self.assertEqual(rotations[f"{Path('gcn.pdf').resolve()}::2"], 180)
        self.assertNotIn(f"{Path('gcn.pdf').resolve()}::3", rotations)

    def test_apply_gemini_page_metadata_accepts_legacy_dict(self) -> None:
        extraction = _sample_extraction().model_copy(update={"page_metadata": {"1": 270}})
        state: dict[str, object] = {}
        with patch("views.entry_actions.st.session_state", state):
            updated = apply_gemini_page_metadata_to_viewer(Path("gcn.pdf"), extraction)
            rotations = dict(state["page_rotations"])

        self.assertTrue(updated)
        self.assertEqual(rotations[f"{Path('gcn.pdf').resolve()}::1"], 270)

    def test_run_ocr_extraction_uses_openai_and_applies_form(self) -> None:
        extraction = object()
        remember_ai_config = Mock()
        with (
            patch("views.entry_actions.extract_land_certificate_with_gemini") as gemini,
            patch("views.entry_actions.extract_land_certificate", return_value=extraction) as openai,
            patch("views.entry_actions.apply_extraction_to_form") as apply_form,
        ):
            result = run_ocr_extraction(
                preview_path=Path("gcn.png"),
                ai_provider="OpenAI",
                api_key="secret",
                model="openai-test",
                remember_ai_config=remember_ai_config,
            )

        self.assertIs(result, extraction)
        gemini.assert_not_called()
        openai.assert_called_once_with(Path("gcn.png"), api_key="secret", model="openai-test")
        apply_form.assert_called_once_with(extraction)
        remember_ai_config.assert_called_once()

    def test_ocr_full_document_prefills_form_then_saves_case_to_sqlite(self) -> None:
        extraction = _sample_extraction()
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "cases.db"
            init_db(db_path)
            result, form_state, remember_ai_config = _run_mock_gemini_ocr(extraction)

            case_id = create_case(
                db_path,
                {
                    "customer_type": "individual",
                    "case_status": DEFAULT_CASE_STATUS,
                    "execution_month": datetime.now().strftime("%m/%Y"),
                    "payment_status": UNPAID_STATUS,
                    "contract_number": "HD-OCR-001",
                    "asset_description": form_state["entry_asset_description"],
                    "customer_info": form_state["entry_customer_info"],
                    "customer_address": form_state["entry_customer_address"],
                    "citizen_id": form_state["entry_citizen_id"],
                    "personal_note": form_state["entry_personal_note"],
                    "so_thua_dat": form_state["so_thua"],
                    "so_to_ban_do": form_state["so_to"],
                    "dia_chi_thua_dat": form_state["land_address"],
                    "owner_name": form_state["owner_name"],
                    "valuation_fee_number": "1.200.000",
                },
            )
            saved = get_case(db_path, case_id)

        self.assertIs(result, extraction)
        remember_ai_config.assert_called_once()
        self.assertIsNotNone(saved)
        assert saved is not None
        self.assertEqual(saved["contract_number"], "HD-OCR-001")
        self.assertEqual(saved["payment_status"], UNPAID_STATUS)
        self.assertEqual(saved["customer_info"], "Nguyen Van A")
        self.assertEqual(saved["customer_address"], "12 Le Loi")
        self.assertEqual(saved["citizen_id"], "012345678901")
        self.assertEqual(saved["so_thua_dat"], "123")
        self.assertEqual(saved["so_to_ban_do"], "45")
        self.assertIn("Phuong Tan Loi", saved["asset_description"])
        self.assertEqual(saved["valuation_fee_number"], 1200000)

    def test_export_excel_after_ocr_prefills_form_values(self) -> None:
        extraction = _sample_extraction()
        result, form_state, remember_ai_config = _run_mock_gemini_ocr(extraction)
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            template_path = root / "form_template.xlsx"
            output_path = root / "form_output.xlsx"
            _create_excel_template(template_path)

            fill_template(
                template_path,
                output_path,
                {
                    "contract_number": "HD-OCR-EXCEL",
                    "asset_description": form_state["entry_asset_description"],
                    "customer_info": form_state["entry_customer_info"],
                    "customer_address": form_state["entry_customer_address"],
                    "citizen_id": form_state["entry_citizen_id"],
                    "personal_note": form_state["entry_personal_note"],
                    "valuation_fee_number": "1.200.000",
                },
            )
            workbook = load_workbook(output_path, data_only=True)
            worksheet = workbook["Up Hs"]
            values = {
                "contract_number": worksheet["D2"].value,
                "asset_description": worksheet["D5"].value,
                "customer_info": worksheet["D11"].value,
                "valuation_fee_number": worksheet["E13"].value,
                "customer_address": worksheet["D23"].value,
                "citizen_id": worksheet["D24"].value,
                "personal_note": worksheet["D25"].value,
            }
            workbook.close()

        self.assertIs(result, extraction)
        remember_ai_config.assert_called_once()
        self.assertEqual(values["contract_number"], "HD-OCR-EXCEL")
        self.assertIn("Thửa đất số 123", values["asset_description"])
        self.assertIn("tờ bản đồ số 45", values["asset_description"])
        self.assertEqual(values["customer_info"], "Nguyen Van A")
        self.assertEqual(values["valuation_fee_number"], 1200000)
        self.assertEqual(values["customer_address"], "12 Le Loi")
        self.assertEqual(values["citizen_id"], "012345678901")
        self.assertEqual(values["personal_note"], "OCR toan bo tai lieu")

    def test_save_original_pdf_and_image_into_case_folder(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_pdf = root / "gcn.pdf"
            source_image = root / "gcn.jpg"
            source_pdf.write_bytes(b"%PDF-1.4\nsample")
            source_image.write_bytes(b"\xff\xd8\xff sample")

            folder = case_folder(
                root / "cases",
                case_id=12,
                contract_number="HD/2026:001",
                customer_name="Nguyen Van A - 0905222125 (Thu tien)",
            )
            saved_pdf = save_original_file(source_pdf, "Giay chung nhan.pdf", folder)
            saved_image = save_original_file(source_image, "Anh GCN.jpg", folder)
            self.assertIsNotNone(saved_pdf)
            self.assertIsNotNone(saved_image)
            assert saved_pdf is not None
            assert saved_image is not None
            self.assertTrue(saved_pdf.exists())
            self.assertTrue(saved_image.exists())
            self.assertEqual(saved_pdf.read_bytes(), b"%PDF-1.4\nsample")
            self.assertEqual(saved_image.read_bytes(), b"\xff\xd8\xff sample")
            self.assertEqual(saved_pdf.parent.name, "originals")
            self.assertEqual(saved_image.parent.name, "originals")
            self.assertEqual(saved_pdf.parent.parent.name, "HD-2026-001 - Nguyen Van A")
            self.assertNotIn("/", saved_pdf.parent.parent.name)
            self.assertNotIn(":", saved_pdf.parent.parent.name)

    def test_save_original_file_does_not_overwrite_duplicate_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            first_source = root / "first.pdf"
            second_source = root / "second.pdf"
            first_source.write_bytes(b"first")
            second_source.write_bytes(b"second")
            folder = case_folder(root / "cases", case_id=13, contract_number="HD-013")

            first_saved = save_original_file(first_source, "GCN.pdf", folder)
            second_saved = save_original_file(second_source, "GCN.pdf", folder)

            self.assertIsNotNone(first_saved)
            self.assertIsNotNone(second_saved)
            assert first_saved is not None
            assert second_saved is not None
            self.assertNotEqual(first_saved, second_saved)
            self.assertEqual(first_saved.name, "GCN.pdf")
            self.assertEqual(second_saved.name, "GCN_2.pdf")
            self.assertEqual(first_saved.read_bytes(), b"first")
            self.assertEqual(second_saved.read_bytes(), b"second")


if __name__ == "__main__":
    unittest.main()
